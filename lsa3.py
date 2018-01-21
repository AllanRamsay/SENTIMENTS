#!/usr/bin/python
# -*- coding: utf-8 -*-
#08/01/18   taken from PAR 
#15/01/18   added tokeniser,emoji finder,
#           codaformat support,
#           support for multiple training sets
 
import re, os, sys, shutil, scipy, numpy, codecs
from numpy import dot

import LSA.a2bw
import json
from LSA.useful import *
from nltk.tokenize import sent_tokenize, word_tokenize
import LSA.stemmer

try:
    import dtw
except:
    print "I don't think we currently need dtw"

from openpyxl import Workbook
from openpyxl.styles import colors, Font, Color

def CellColourCheck(i, dc, sColour, sCell):
    try:
        if str(i).find("\t" + sColour)>-1:
            dc.append(sCell)
    except:
        pass

    try:
        i = i.replace("\t"+sColour,"")
    except:
        pass

    return i

def ReplaceColours(ws, dc, oCol):
    for sCell in dc:
        ws[sCell].style = oCol

def toXLSX(lines, xlsxfile="test.xlsx"):
    global moMyArgs
    global mBlackText
    global mRedText
    global mBlueText
    global mGreenText
    global mPurpleText

    # swap the keys and values
    dcEMOJIS = {}
    for k,v in EMOJIS.iteritems():
        dcEMOJIS[v] = k

    dcRed = []
    dcBlue = []
    dcBlack = []
    dcPurple = []
    dcGreen = []

    print "in toXLSX, writing file", xlsxfile

    wb = Workbook()
    ws = wb.active
    try:
        lines = lines.split("\n")
    except:
        pass
    
    nRow=1
    for l in lines:
        lNew = []
        nCol=65
        for i in l:

            for k,v in dcEMOJIS.iteritems():
                try:
                   i = i.replace(k,v)
                except:
                    pass

            sCell = chr(nCol) + str(nRow)

            i = CellColourCheck(i, dcRed, "RED", sCell)
            i = CellColourCheck(i, dcBlack, "BLACK", sCell)
            i = CellColourCheck(i, dcGreen, "GREEN", sCell)
            i = CellColourCheck(i, dcBlue, "BLUE", sCell)
            i = CellColourCheck(i, dcPurple, "PURPLE", sCell)

            lNew.append(i)

            nCol+=1

        if moMyArgs.PRINT_DEV:
            print nRow
            print l
            print lNew
            print ""

        ws.append(lNew)
        nRow+=1

    ReplaceColours(ws, dcBlack, mBlackText)
    ReplaceColours(ws, dcRed, mRedText)
    ReplaceColours(ws, dcBlue, mBlueText)
    ReplaceColours(ws, dcGreen, mGreenText)
    ReplaceColours(ws, dcPurple, mPurpleText)

    wb.save(xlsxfile)

"""
>>> etest = makeTFIDF(ENGLISHDEV)
>>> etraining = makeTFIDF(ENGLISHTRAINING)

# this is just the raw counts
>>> esentidict = makeBaseClassifier(etraining)

# this is the one that produces a basic classifier
>>> k0 = makeScoreDict(training, sentidict, N=2500)
>>> plot(test, k0)

This will generate a list of classifiers with a range of thresholds.

>>> eclassifiers = createAndTuneRange(etraining, etest, esentidict)

# remove misleading words, set a threshold for each 
# sentiment. This a few seconds. The threshold set is now
# stored in the new classifier

>>> k1 = tune(training, k0)

# Put each tweet in a nice looking file: j is the Jaccard value: should be 0.415, f should be 0.586

>>> p, r, f, j, errors, lines = scoreTweets(test, k1, threshold=k1.threshold, out="xxx.xlsx")

"""

TWEETS = 'twitter_samples/tweets.20150430-223406.json'

BASEPATH = 'Data/SAIF/SEMEVAL_2018_CLASSIFICATION/Data/'
BASEPATH = '../Raw/'

global gFiles ; gFiles = {}

gFiles['ENGLISHTRAINING'] = BASEPATH + 'EN/2018-E-c-En-train.txt'
gFiles['ENGLISHDEV'] = BASEPATH + 'EN/2018-E-c-En-dev.txt'
gFiles['ENGLISHTEST'] = BASEPATH + 'EN/2018-E-c-En-test.txt'
gFiles['ENGLISHTWEETSBYID'] = BASEPATH + 'EN/tweetsbyid.txt'
gFiles['ENGLISHREG'] = BASEPATH + 'EN/ALL-EI-reg-En.txt'
gFiles['ENGLISHOC'] = BASEPATH + 'EN/ALL-EI-oc-En.txt'

gFiles['ARABICDEV'] = BASEPATH + 'AR/2018-E-c-Ar-dev.txt'
gFiles['ARABICTRAINING'] = BASEPATH + 'AR/2018-E-c-Ar-train.txt'
gFiles['ARABICTEST'] = BASEPATH + 'AR/2018-E-c-Ar-test.txt'

"""
ENGLISHTRAINING = "Data/SAIF/SEMEVAL_2018_CLASSIFICATION/Data/EN/2018-E-c-En-train.txt"
ENGLISHDEV = "Data/SAIF/SEMEVAL_2018_CLASSIFICATION/Data/EN/2018-E-c-En-dev.txt"
ENGLISHTEST = "Data/SAIF/SEMEVAL_2018_CLASSIFICATION/Data/EN/2018-E-c-En-test.txt"
ENGLISHTWEETSBYID = "Data/SAIF/SEMEVAL_2018_CLASSIFICATION/Data/EN/tweetsbyid.txt"

ARABICDEV = "Data/SAIF/SEMEVAL_2018_CLASSIFICATION/Data/AR/2018-E-c-Ar-dev.txt"
ARABICTRAINING = "Data/SAIF/SEMEVAL_2018_CLASSIFICATION/Data/AR/2018-E-c-Ar-train.txt"
ARABICTEST = "Data/SAIF/SEMEVAL_2018_CLASSIFICATION/Data/AR/2018-E-c-Ar-test.txt"
"""

languagePattern = re.compile("%s(?P<language>.*)/.*"%(BASEPATH))

null = None
false = False
true = True

prefixes = re.compile("\s*((rt\s*)?@\S+)+")
urls = re.compile("//t.co/\S*|http|#|:")

class TWEET:

    def __init__(self, id=False, src=False, text=False, features=False, scores=False, tokens=False):
        self.id = id
        self.features = normalise(features)
        self.src = src
        self.text = text
        self.GS = scores
        self.tokens = tokens
        
    def __repr__(self):
        return self.src

    def normalise(v):
        t = sum(v.values())
        for k in v:
            v[k] = v[k]/t
        return v
morepunc = re.compile("(?P<punc>/|\.+|'|,|!|#|\$|%|\d+|&\S*|\(|\)|\*|\[|\]|-|\+|:|;|=|@|'|`|\")")

def tokenise(s, allowweirdchars=False):
    tokens = []
    for t in word_tokenize(morepunc.sub(" ", s)):
        try:
            t.encode("utf-8")
            tokens.append(t)
        except:
            print "T", t
            if not allowweirdchars:
                continue
            t = t.decode('ISO-8859-1')
            tokens.append(t)
    return tokens

class TESTSET:

    def __init__(self, cols, tweets, idf):
        self.cols = cols
        self.tweets = tweets
        self.idf = idf
        self.words = [w[0] for w in reversed(sortTable(idf))]
        self.makeIndex()

    def __iter__(self):
        for tweet in self.tweets:
            yield tweet

    def __getitem__(self, i):
        return self.tweets[i]

    def makeIndex(self):
        self.index = {}
        for tweet in self.tweets:
            for token in tweet.tokens:
                if not token in self.index:
                    self.index[token] = {}
                for col, score in zip(self.cols, tweet.GS):
                    if score == "1":
                        try:
                            self.index[token][col].append(tweet)
                        except:
                            self.index[token][col] = [tweet]

"""
Remove rubbish from tweets, remove tweets that are very similar to one another
"""

SPACE = re.compile(" +")
def splitEmojis(s0):
    s1 = ""
    for c in s0:
        if ord(c) > 256:
            s1 += " "+c+" "
        else:
            s1 += c
    return SPACE.sub(" ", s1)

EMOJIS = {}
EMOJICOUNT = 0

def isEmoji(t):
    return ord(t[0]) > 3000

def fixEmojis(tweet0):
    tweet1 = []
    global EMOJICOUNT
    global EMOJIS
    for t in tweet0:
        if len(t) > 0:
            if isEmoji(t):
                try:
                    t = EMOJIS[t]
                except:
                    EMOJIS[t] = "eMJ%s"%(EMOJICOUNT)
                    EMOJICOUNT += 1
                    t = EMOJIS[t]
            tweet1.append(t)
    return tweet1

def makeTweet(tweet):
    try:
        tss = tweet.strip().split("\t")
        tweet, src, scores = tss[0], tss[1], tss[2:]
    except:
        tweet, src, scores = "DUMMY", tweet, False
    tweet = tweet.strip().replace("=", "").replace("\\n", " ")
    if LANGUAGE == "AR":
        tweet = SPACE.sub(" ", re.compile(u"_|#|\.+|;|!|؛|!|:|.").sub(" ", LSA.a2bw.convert(tweet, LSA.a2bw.a2bwtable)))

    #t this puts spaces around emojis that are next to each other
    # i didnt do this in mine so dont do it?
    #tweet = splitEmojis(tweet)

    d = {}
    text = prefixes.sub("", src).strip()
    text = t_ProcessTweet(text)
    if LANGUAGE == "AR":
        text = SPACE.sub(" ", replaceAll(text, [(u"؛", " ")]))
        tokens = LSA.stemmer.stemAll(text) 
    else:
        if moMyArgs.Tokeniser=="T":
            tokens = LSA.NLTKTokeniser2.casual_tokenize(text.strip())
        else:
            tokens = [w for w in tokenise(prefixes.sub("", text.lower()).strip())]
    tokens = [token for token in fixEmojis(tokens) if not token == ""]

    if moMyArgs.PRINT_DEV:
        print "---tokenised",tweet,"---"
        for r in tokens:
            print r,
        print ""

    for w in tokens:
        try:
            d[w] += 1
        except:
            d[w] = 1
    return TWEET(id=tweet, src=src, text=text, features=d, scores=scores, tokens=tokens)

    #t lDocs is a list of docs for training
def makeTFIDF(lDocs, threshold=0, N=False):
    global LANGUAGE
    LANGUAGE = languagePattern.match(lDocs[0]).group("language")
    idf = {}
    cols = False
    tweets = []
    index = {}
    """
    just read them and add them to a list

    calculate the tf vector and accumulate the idf vector as you go
    """
    for docs in lDocs:
        cols = False
        print "\nmakeTFIDF with file: " ,docs
        for tweet in codecs.open(docs, encoding="UTF-8").read().split("\n"):
            if tweet == "":
                continue
            if not cols:
                cols = tweet.split()[2:]
                continue
            tweets.append(makeTweet(tweet))
            for w in tweets[-1].features:
                try:
                    idf[w] += 1
                except:
                    idf[w] = 1
    """
    remove singletons from the idf count
    """
    for w in idf.keys():
        if idf[w] == 1:
            del idf[w]
        else:
            idf[w] = 1.0/float(idf[w])
    """
    multiply by idf value, remove terms that aren't in the idfvector
    (since they were singletons) or that have very low scores as you go
    """
    for tweet in tweets:
        for w in tweet.features.keys():
            try:
                t = float(tweet.features[w])*idf[w]
                if t < threshold:
                    del tweet.features[w]
                else:
                    tweet.features[w] = t
            except:
                del tweet.features[w]
        """
        If we've set a maximum number of features/tweet, set things so
        that's how many we've got
        """
        if N and len(d.features) > N:
            bestN = set([x[0] for x in sortTable(d.features)][:N])
            for f in d.features.keys():
                if not f in bestN:
                    del d.features[f]
        else:
            bestN = False
    """
    re-prune the idf table. Not sure that that's actually going to do anything!
    """
    if threshold > 0:
        allgoodtf = set()
        for d in tweets:
            for f in d.features:
                allgoodtf.add(f)
        for f in idf.keys():
            if not f in allgoodtf:
                del idf[f]
    print "%s tweets in set: "%(str(len(tweets)))
    return TESTSET(cols, tweets, idf)
        
def makeBaseClassifier(training, N=sys.maxint, printing=False):
    invemojis = {EMOJIS[x]:x for x in EMOJIS}
    sdict = [{} for i in range(len(training.cols))]
    for tweet in training.tweets:
        if printing:
            print tweet.tokens
            print tweet.GS
        for w in tweet.tokens:
            for score, scorecard in zip(tweet.GS, sdict):
                s = 2 if w in invemojis else 1
                if score == '1':
                    try: 
                        scorecard[w] += s
                    except:
                        scorecard[w] = s
                    try: 
                        scorecard["%%%%%"] += s
                    except:
                        scorecard["%%%%%"] = s
        N -= 1
        if N == 0:
            break
    return CLASSIFIER(scorelist=sdict, columns=training.cols)

def average(scores):
    t = sum(scores)
    for i, x in enumerate(scores):
        scores[i] = x/t
    return scores

def joinscores(scores):
    return "\t".join(["%.2f"%(s) for s in scores])

def getScores(word, sdict, printing=False, dicttype="FULL"):
    scores = []
    rawcount = []
    for d in sdict.scorelist:
        try:
            """
            d[word] is number of tweets that express S that contain word
            d["%%%%%"] is total number of words in tweets that express S

            So d[word]/d["%%%%"] is just p(W|S), and I'm sort of using
            it as a substitute for p(S|W). The reason it works is
            because I use it for calculating the *relative* likelihood
            that W will express S. If "ratbag" is more likely to
            express anger than it is to express joy then I will let it
            vote for anger and against joy.
            """
            score = float(d[word])/float(d["%%%%%"])
            scores.append(score)
            rawcount.append(float(d[word]))
        except:
            rawcount.append(0.0)
            scores.append(0.0)
    """
    Removing the last column, which happens to be "trust" produces an
    overall improvement. This needs investigation
    """
    # scores = scores[:-1]
    """
    Just normalise them. Note that this wipes out any record of how
    common the sentiment was -- we are just looking at the relative
    likelihoods of this word occuring with different sentiments. We
    will recover information about how likely a given sentiment is
    when we set the local thresholds -- if 10% of tweets are angry,
    then choosing a threshold for anger that marks 10% of tweets as
    being angry will deal with what we know about the absolute
    likelihood of a tweet being angry.
    """
    scores = average(scores)
    if printing:
        print "rawcounts"
        print joinscores(rawcount)
        print "probability that a tweet that expresses sentiment I contains '%s'"%(word)
        print joinscores(scores)
    if dicttype == "TFIDF":
        return [word]+scores
    """
    Subtract the average score over all sentiments for this word from
    the individual scores for each sentiment. This is a sort of local
    IDF step -- if a word is equally common for all sentiments then we
    don't want it vote for any of them, and if it is below the overall
    average for some sentiment we want it to be allowed to vote
    *against* it.
    """
    a = sum(scores)/len(scores)
    for i, x in enumerate(scores):
        scores[i] = x-a
    if printing:
        print "distance from the mean for each sentiment"
        print joinscores(scores)
    if dicttype == "DIST":
        return [word]+scores
    """
    Multiply weights by standard deviation. This gives extra weight to
    words that have very skewed distributions. I can't really give a
    coherent justification for this, and indeed I can't give a
    reasoned explanation of whether we should use sum(map(abs(x-a), scores)) 
    or sqrt(sum(map(x-a)**2, scores))) (the latter is the actual standard
    deviation, but sqrt(sum(scores))) works better)
    """
    sd = sqrt(sum(map(lambda x: (x-a)**2, scores)))/len(scores)*1000
    for i, x in enumerate(scores):
        scores[i] = x*sd
    if printing:
        print "multiply that by standard deviation"
        print joinscores(scores)
    return [word]+scores

def scoredict2scorelist(scoredict):
    scorelist = [{} for x in scoredict.values()[0]]
    for w in scoredict:
        for i, v in enumerate(scoredict[w]):
            scorelist[i][w] = v
    return scorelist

def scorelist2scoredict(scorelist):
    scoredict = {}
    for col in scorelist:
        for w in col:
            try:
                scoredict[w].append(col[w])
            except:
                scoredict[w] = [col[w]]
    return scoredict

class CLASSIFIER:

    def __init__(self, scorelist=False, scoredict=False, threshold=0.3, columns=False):
        self.scorelist = scorelist
        self.scoredict = scoredict
        if scorelist and not scoredict:
            self.scoredict = scorelist2scoredict(scorelist)
        if scoredict and not scorelist:
            self.scorelist = scoredict2scorelist(scoredict)
        self.threshold = threshold
        self.columns = columns

    def show(self, out=sys.stdout):
        with safeout(out) as write:
            for w in self.scoredict:
                write("%s\t%s\n"%(w, joinscores(self.scoredict[w])))
            
def makeClassifier(training, classifier, N=sys.maxint, dicttype="FULL"):
    words = [x[0] for x in reversed(sortTable(training.idf))][:N]
    scores = []
    scoredict = {}
    for word in words:
        try:
            scores.append(getScores(word, classifier, dicttype=dicttype))
            scoredict[word] = scores[-1][1:]
        except:
            pass
    return CLASSIFIER(scoredict=scoredict, columns=training.cols)
        
def showAllScores(allscores, out=sys.stdout):
    with safeout(out) as write:
        write("\tanger	anticipation	disgust	fear	joy	love	optimism	pessimism	sadness	surprise	trust\n")
        for r in allscores:
            print r
            write("%s\t%s\n"%(r[0], joinscores(r[1:])))

def scoreTweet(tweet, classifier, threshold=0.5, sFrom=""):
    if sFrom!="" and False:
        print "in scoreTweet from", sFrom, "t=", str(threshold)
    scoredict = classifier.scoredict
    sentiments = [0 for x in classifier.columns]
    scores = [0 for x in classifier.columns]
    best = [(0, " ") for x in classifier.columns]
    worst = [(sys.maxint, " ") for x in classifier.columns]
    for word in tweet.tokens:
        try:
            for i, s in enumerate(scoredict[word]):
                sentiments[i] += s
                if s > best[i][0]:
                    best[i] = (s, word)
                if s < worst[i][0]:
                    worst[i] = (s, word)
        except:
            pass
    if threshold or threshold==0: #t allow a threshold of 0!
        m = max(sentiments)
        if m > 0:
            for i, x in enumerate(sentiments):
                scores[i] = x/m
                try:
                    lt = threshold[i]
                except:
                    lt = threshold
                if scores[i] < lt:
                    sentiments[i] = 0
                else:
                    sentiments[i] = 1
    return sentiments, scores, best, worst

def prfj(right, wrong, missing):
    try:
        right = float(right)
        p = right/(right+wrong)
        r = right/(right+missing)
        f = 2*p*r/(p+r)
        return (p, r, f, right/(right+wrong+missing))
    except:
        return (0, 0, 0, 0)

M11 = 0
M10 = 1 # wrong
M01 = 2 # missing
M00 = 3

def convertible(w):
    if LANGUAGE == "AR" and ord(w[0]) < 256:
        return LSA.a2bw.convert(w, LSA.a2bw.bw2atable)
    else:
        return w

def codaformat(testset, classifier, out=sys.stdout, threshold=None, singleColumn=None):
    if threshold == None:
        threshold = classifier.threshold    
    J = [[0]*len(testset.cols) for i in range(4)]
    lines = []
    worderrors = [{} for x in testset.cols]
    with safeout(out, encoding="UTF-8") as write:
        write("ID\tTweet\tanger\tanticipation\tdisgust\tfear\tjoy\tlove\toptimism\tpessimism\t sadness\tsurprise\ttrust\n")
        for I, tweet in enumerate(testset.tweets):
            cols = testset.cols
            sentiments, scores, best, worst = scoreTweet(tweet, classifier, threshold=threshold)
            write("%s\t%s\t%s\n"%(tweet.id, tweet.src, "\t".join(map(str, sentiments))))
        
def scoreTweets(testset, classifier, out=False, threshold=null, singleColumn=None, sFrom=""):
    if threshold == null:
        threshold = classifier.threshold    
    J = [[0]*len(testset.cols) for i in range(4)]
    lines = []
    worderrors = [{} for x in testset.cols]
    for I, tweet in enumerate(testset.tweets):
        cols = testset.cols

        sFrom2=""
        if (sFrom=="createAndTune"): #t
            sFrom2="scoreTweets"
            DUMMY=1

        sentiments, scores, best, worst = scoreTweet(tweet, classifier, threshold=threshold, sFrom=sFrom2)

        for n, (col, sentiment, score, gs) in enumerate(zip(cols, sentiments, scores, map(int, tweet.GS))):
            if not singleColumn == None and not n == singleColumn:
                continue
            if sentiment == 1:
                if gs == 1:
                    J[M11][n] += 1
                    for token in tweet.tokens:
                        try:
                            worderrors[n][token] += 1
                        except:
                            worderrors[n][token] = 1
                    sentiments[n] = "%s:%.2f:\t%s"%(col, score, "GREEN")
                else:
                    J[M10][n] += 1
                    for token in tweet.tokens:
                        try:
                            worderrors[n][token] -= 1
                        except:
                            worderrors[n][token] = -1
                    sentiments[n] = "%s:%.2f\t%s"%(col, score, "RED")
            else:
                if gs == 1:
                    J[M01][n] += 1
                    sentiments[n] = "%s:%.2f\t%s"%(col, score, "BLUE")
                else:
                    J[M00][n] += 1
            tokens = " ".join(tweet.tokens)
            if LANGUAGE == "AR":
                tokens = LSA.a2bw.convert(tokens, LSA.a2bw.bw2atable)
        if singleColumn == None:
            sentiments.append(tokens)
        try:
            if singleColumn == None:
                lines.append(sentiments)
                lines.append(["%s:%.2f"%(convertible(b[1]), b[0]) for b in best])
                lines.append(["%s:%.2f"%(convertible(w[1]), w[0]) for w in worst])
                lines.append([""]*len(best))
        except Exception as e:
            print "Couldn't convert %s (tweet %s)"%(tweet.text, I)
            print "BEST", best
            print "WORST", worst
            raise(e)
    if singleColumn == None:
        localJ = []
        for i in range(len(J[M11])):
            j = [j[i] for j in J]
            localJ.append("%.2f, %.2f"%tuple(prfj(j[M11], j[M10], j[M01])[-2:]))
        lines.append(cols)
        lines.append(localJ)
    # print "\t".join(["%.2f"%(prfj(r, w, m)[-1]) for r, w, m in zip(J[M11], J[M10], J[M01])])
    if out:
        toXLSX(lines, xlsxfile="%s.xlsx"%(out))
        writecsv(lines, out=codecs.open("%s.csv"%(out), encoding="UTF-8", mode="w"))
    score = prfj(float(sum(J[M11])), float(sum(J[M10])), float(sum(J[M01])))
    return score+(worderrors, lines)

"""
This is either yuk or clever, depending on your point of view. It is
in some ways like Brill retagging -- look at things that you are
getting wrong and change your representation. I was worried that it
might actually involve a huge percentage of words, which would mean
that the basic method was rubbish and this was doing all the work, but
the numbers are quite small -- around 5% for English, 1% for Arabic --
so I'm happy that it's a corrective for a method that is basically OK
but makes a moderate number of errors.

Each word has a score associated with it for every sentiment. This is
supposed to be a measure of how much evidence this word supplies for
this sentiment. But it might happen that although this word appears to
provide evidence for a sentiment, actually more tweets that contain
this word don't express this sentiment. So we do a check here -- if W
appears to want to vote for S, but most tweets that contain W don't
express S, then we give W a negative score for S (used to make it 0,
but actually using a negative score is better). It would be good to
calculate the right negative score for each W:S combination: at the
moment I'm just using an experimentally determined value (er,
guessing).

"""
def selfcorrect(training, classifier, N=1, threshold=0.3):
    removed = 0
    allwords = 0
    for i in range(N):
        (p, r, f, j, worderrors, lines) = scoreTweets(training, classifier, out=False, threshold=threshold, sFrom="selfcorrect")
        for sd, wd in zip(classifier.scorelist, worderrors):
            for w in wd:
                allwords += 1
                """
                Change the opinion of a stupid word if it thought it was being helpful
                """
                if wd[w] < 0 and w in sd and sd[w] > 0:
                    # print "%s: %.2f, %s"%(w, sd[w], joinscores(classifier.scoredict[w]))
                    sd[w] = -2
                    removed += 1
        classifier = CLASSIFIER(scorelist=classifier.scorelist, columns=classifier.columns)
    print "  altered %s out of %s words (%.1f%%)"%(removed, allwords, 100*float(removed)/float(allwords))
    return classifier

def setThresholds(training, scoredict):
    return [plot(training, scoredict, singleColumn=i, printing=False) for i in range(len(training.cols))]

"""
tuning involves getting rid of words that have led to bad decisions
and then finding the best per-column thresholds. The first bit in turn
depends on the threshold which was set for using the initial
classifier, so you have to optimise for that as well.
"""
def tune(training, scoredict, threshold=0.15, N=3):
    scoredict.threshold = threshold
    print "tune %.1f"%(threshold)
    for i in range(N):
        print "  selfcorrect, round %s"%(i)
        scoredict = selfcorrect(training, scoredict, threshold=scoredict.threshold)
        print "  setThresholds, round %s"%(i)
        scoredict.threshold = setThresholds(training, scoredict)
    return scoredict

def createAndTune(training, test, classifier, threshold=0.15, N1=2500, N2=3):
    k = makeClassifier(training, classifier, N=N1)
    k = tune(training, k, threshold=threshold, N=N2)
    p, r, f, j, errors, lines = scoreTweets(test, k, sFrom="createAndTune")
    print "T %.2f, P %.3f, R %.3f, Jaccard %.3f F0 %.3f"%(threshold, p, r, j, f)
    """
    add the jaccard score to the classifier so we can compare classifiers to choose the best
    """
    k.jaccard = j
    return k

def createAndTuneRange(training, test, t=1, end=0, step=0.1, N1=2500, N2=1):
    classifiers = []
    invemojis = {EMOJIS[x]:x for x in EMOJIS}
    baseclassifier = makeBaseClassifier(training)
    print "Start training: %s to %s"%(t, end)
    while t >= end:
        classifiers.append(createAndTune(training, test, baseclassifier, threshold=t, N1=N1, N2=N2))
        t -= step
    return classifiers

def plot(testset, scoredict, threshold=1.0, end=0, step=0.1, singleColumn=None, printing=True):
    if not singleColumn == None:
        print "    setting threshold for %s"%(testset.cols[singleColumn])
    m = []
    while threshold > end:
        (p, r, f, j, w, l) = scoreTweets(testset, scoredict, threshold=threshold, out=False, singleColumn=singleColumn, sFrom="plot")
        m.append((threshold, p, r, f, j))
        if printing:
            print "%.1f\t%.2f\t%.2f\t%.3f\t%.3f"%(m[-1])
        threshold -= step
    lastT = 1.0
    bestJ = 0
    bestJt = 1.0
    allJ = []
    sweetspot = False
    for threshold, p, r, f, j in m:
        if j > bestJ:
            bestJ = j
            bestJt = threshold
        allJ.append((j, threshold))
        if not sweetspot and r > p:
            sweetspot = threshold+(lastT-threshold)/2.0
        lastT = threshold
    m.append((sweetspot,)+scoreTweets(testset, scoredict, threshold=sweetspot, out=False, sFrom="plot(2)")[:-1])
    if printing:
        print "--------------------------"
        print "%.1f\t%.2f\t%.2f\t%.3f\t%.3f"%(m[-1])
        print "bestJ %.2f %.3f"%(bestJt, bestJ)
    if not singleColumn == None:
        return bestJt

def export(d, training=False, out=sys.stdout):
    try:
        d = d.scoredict
    except:
        pass
    invemojis = {EMOJIS[x]:x for x in EMOJIS}
    with safeout(out) as write:
        try:
            write("\t".join([""]+training.cols)+"\n")
        except:
            pass
        for w in sorted(d.keys()):
            try:
                k = invemojis[w]
            except:
                if LANGUAGE == "AR":
                    k = w.split(":")
                    r = LSA.a2bw.convert(k[0], LSA.a2bw.bw2atable)
                    if len(k) > 1:
                        k = r+":%s"%(k[1])
                    else:
                        k = r
                else:
                    k = w
            write("\t".join([k]+["%.3f"%(s) for s in d[w]])+"\n")

 
#t --------------------------------------------my stuff--------------------------------------------
#t --------------------------------------------my stuff--------------------------------------------
#t --------------------------------------------my stuff--------------------------------------------

import nabeela_common
from LSA.TNAEmojiHelper import *
import LSA.NLTKTokeniser2
from LSA.ANRVStemmer import *	# Englsh 4 way stemmer

try:
    from LSA.TNAWordAnalyser import *
except Exception as e:
    print e

from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell
from openpyxl.styles import NamedStyle, Font, Border, Side

global dcNONEmojiIndex ; dcNONEmojiIndex = {} # non emojis
global dcEmojiIDIndex ; dcEmojiIDIndex = {} # emoji replacements, the index is the ID

global mRedFill ; mRedFill = PatternFill(start_color='FFFF0000',end_color='FFFF0000', fill_type='solid')
global mRedText ; mRedText = NamedStyle(name="red") ; mRedText.font = Font(color="ff0000")
global mBlueText ; mBlueText = NamedStyle(name="blue") ; mBlueText.font = Font(color="00ff00")
global mGreenText ; mGreenText = NamedStyle(name="green") ; mGreenText.font = Font(color="0000ff")
global mBlackText ; mBlackText = NamedStyle(name="black") ; mBlackText.font = Font(color="000000")
global mPurpleText ; mPurpleText = NamedStyle(name="purple") ; mPurpleText.font = Font(color="FF00FF")

class MyArgs(nabeela_common.MyArgsBase):
	Tokeniser = ""
	ReplaceEmojis = False
	t = 1.0
	PRINT_DEV = False
	Train = None
	Test = None

global moMyArgs

def t_ProcessTweet(sTweet):
	global dcNONEmojiIndex
	global dcEmojiIDIndex
	global moMyArgs

	bRemoveHashtags = False
	bRemoveHashtagSymbol = False
	bCopyRemoveSplitHashtag = False
	bRemoveHashtagSymbolSplit = False

	if moMyArgs.LowerCase:
		sTweet = sTweet.lower()

	if moMyArgs.RemovePunctuationEN:
		sTweet = RemovePunctuationCharactersEN(sTweet)

	if moMyArgs.ReplaceEmojis==True:
		sTweet = SeparateEmojis(sTweet, dcNONEmojiIndex, EMOJIS, dcEmojiIDIndex)

	if moMyArgs.HashtagProcessMode == "REMOVEHASHTAGS": bRemoveHashtags = True
	if moMyArgs.HashtagProcessMode == "REMOVE#": bRemoveHashtagSymbol = True
	if moMyArgs.HashtagProcessMode == "COPY_REMOVE#_SPLIT": bCopyRemoveSplitHashtag = True
	if moMyArgs.HashtagProcessMode == "REMOVE#_SPLIT": bRemoveHashtagSymbolSplit = True

	if bRemoveHashtags or moMyArgs.RemoveAtTags or moMyArgs.RemoveURLs:
		sTweet = RemoveTokensByClass(sTweet, bRemoveHashtags, False, moMyArgs.RemoveAtTags, moMyArgs.RemoveURLs)

	if bRemoveHashtagSymbol==True:
		if sTweet!=None: sTweet = sTweet.replace("#", "")

	if bRemoveHashtagSymbolSplit==True:
		if sTweet!=None: sTweet = sTweet.replace("#","").replace("_"," ").replace("-"," ").strip()

	if bCopyRemoveSplitHashtag==True:
		sTweet = nabeela_common.ExpandTweetHashtags(sTweet)

	if moMyArgs.Stemmer==4:
		oS = ARNVStemmerHelper()
		oS.Clear()
		oS.Append(sTweet)
		sTweet = oS.StemARNV()

	return sTweet

def t_WriteOut(testset):
	global mRedFill
	global mRedText
	global mBlueText
	global mGreenText
	global mBlackText
	global mPurpleText

	# write out in format we can work with
	oHeader = [u"Dialect.0", u"Dialect.1", u"Dialect.2", u"Emotion.0", u"Emotion.1", u"Polarity", u"Sentence", u"Subdialect.0.0", u"Subdialect.0.1", u"Subdialect.0.2", u"Subdialect.0.3", u"Subdialect.0.4", u"Subdialect.0.5", u"Subdialect.0.6", u"Subdialect.0.7", u"Subdialect.1.0", u"Subdialect.2.0", u"Subdialect.2.1", u"Subdialect.3.0", u"Time.time", u"User.Name", u"Words", u"_id", u"OriginalTweet", u"TESTDATA" ,]

	lNewData = []

	for tweet in testset.tweets:
		sTweet = tweet.text
		lRowData = []
		lRowData.append("") #Dialect.0
		lRowData.append("") #Dialect.1
		lRowData.append("") #Dialect.2
		lEmotions = []
		lRowData.append("") #Emotion.1
		lRowData.append("") #Polarity
		lRowData.append(sTweet) #Sentence
		lRowData.append("") #Subdialect.0.0
		lRowData.append("") #Subdialect.0.1
		lRowData.append("") #Subdialect.0.2
		lRowData.append("") #Subdialect.0.3
		lRowData.append("") #Subdialect.0.4
		lRowData.append("") #Subdialect.0.5
		lRowData.append("") #Subdialect.0.6	
		lRowData.append("") #Subdialect.0.7	
		lRowData.append("") #Subdialect.1.0	
		lRowData.append("") #Subdialect.2.0	
		lRowData.append("") #Subdialect.2.1	
		lRowData.append("") #Subdialect.3.0	
		lRowData.append("") #Time.time	
		lRowData.append("") #User.Name	
		lRowData.append("") #Words	
		lRowData.append(tweet.id) #_id

		lNewData.append(lRowData)

	wb = Workbook()

	sOutputFile = 'lsa2.xlsx'

	oWS = wb.active

	oWS.title = "Sheet1"

	oWS.append(oHeader)

	nColor = 1 
	nRow = 1
	for oRow in lNewData:
		oWS.append(oRow)
		if nColor==1:
			oWS['F' + str(nRow)].style = mRedText
		if nColor==2:
			oWS['F' + str(nRow)].style = mBlueText
		if nColor==3:
			oWS['F' + str(nRow)].style = mGreenText
		nRow+=1
		nColor+=1
		if nColor>3:
			nColor=1

	wb.save(filename = sOutputFile)

def classify(tweet, classifierFile="classifier.pck"):
    global classifier
    global LANGUAGE
    try:
        classifier
    except NameError:
        classifier = load(classifierFile)
    LANGUAGE = classifier.language 
    global moMyArgs
    moMyArgs = classifier.moMyArgs
    TNAWordAnalyser.gdEmoji = classifier.gdEmoji
    sentiments, scores, best, worst = scoreTweet(makeTweet(tweet), classifier)
    return " ".join(classifier.columns[i] for i, s in enumerate(sentiments) if s == 1)

# Standard boilerplate to call the main() function to begin
# the program.
if "lsa3.py" in sys.argv[0]:
    global moMyArgs
    global LANGUAGE

    moMyArgs = MyArgs()
    for sArg in sys.argv:
        sArg = sArg.upper()

        if (sArg.startswith("T=")):
            moMyArgs.t = float(sArg.split("=")[1])
        elif (sArg.startswith("TOKENISER=")):
            moMyArgs.Tokeniser = str(sArg.split("=")[1])
        elif (sArg.startswith("DEV=")):
            moMyArgs.Dev = str(sArg.split("=")[1])
        elif (sArg.startswith("TEST=")):
            moMyArgs.Test = str(sArg.split("=")[1])
        elif (sArg.startswith("TRAIN=")):
            moMyArgs.Train = str(sArg.split("=")[1])
        elif sArg.startswith("REPLACEEMOJIS"):
            moMyArgs.ReplaceEmojis = True
        elif sArg.startswith("DUMP"):
            moMyArgs.Dump = True
        elif (sArg.startswith("PRINT_DEV=")):
            if str(sArg.split("=")[1]).upper()=="Y":
                moMyArgs.PRINT_DEV=True

    nabeela_common.LoadArgs(sys.argv, moMyArgs)
    nabeela_common.PrintArgs(moMyArgs)

    lTrainingData = moMyArgs.Train.split(",")
    allData = [lTrainingData]
    try:
        lDevData = moMyArgs.Dev.split(",")
        allData.append(lDevData)
    except:
        pass
    try:
        lTestData = moMyArgs.Test.split(",")
        allData.append(lTestData)
    except:
        pass

    for lKeys in allData:
        for n, sKey in enumerate(lKeys):
            try:
                lKeys[n] = gFiles[sKey]
            except:
                print "Invalid source: ", sKey
                exit()

    """
    You *always* have to train a set of classifiers. You can't train a
    set, extract the thresholds and then use those in a separate pass,
    because training a classifier does more than just select the
    thresholds.

    For the training part you have to specify TRAIN and DEV. If you
    then want to save a test, you also have to specify TEST. Given
    that doing codaformat requires you to have specified TEST, we
    don't now actually need CODAFORMAT as a flag, because we can just
    try to do a test: if TEST wasn't set then this will throw an
    exception, which we will catch and ignore.

    So to do a round of training without then generating an output file do 

    lsa3.py TOKENISER=T REPLACEEMOJIS LOWERCASE HASHTAGPROCESSMODE=COPY_REMOVE#_SPLIT STEMMER=4 TRAIN=ENGLISHTRAINING DEV=ENGLISHDEV 

    If you want to generate an output file as well do
    
    lsa3.py TOKENISER=T REPLACEEMOJIS LOWERCASE HASHTAGPROCESSMODE=COPY_REMOVE#_SPLIT STEMMER=4 TRAIN=ENGLISHTRAINING DEV=ENGLISHDEV TEST=ENGLISHTEST

    (note no argument CODAFORMAT)
    ...
    
    4. to see if getting tweets by ID gives us better thresholds
    ENGLISHTWEETSBYID, ENGLISHDEV

    5. to see if the tweets from the OC and REG tasks help (only cover 4 emotions though...)
    ENGLISHTRAINING,ENGLISHREG,ENGLISHOC ENGLISHDEV
    """

    print "TRAINING: ", lTrainingData
    print "DEV: ", lDevData
    try:
        print "TEST: ", lTestData
    except:
        print "NO TEST DATA SUPPLIED"
    
    training = makeTFIDF(lTrainingData)
    dev = makeTFIDF(lDevData)

    #Whether or not you want to do a submittable test, you have to train a set of classifiers first
    k = createAndTuneRange(training, dev, t=moMyArgs.t, step=0.1)
    bestj = 0
    besti = -1
    for i, c in enumerate(k):
        if c.jaccard > bestj:
            bestj = c.jaccard
            best = c
            besti = i
            best.language = LANGUAGE
    print "best classifier is %s"%(besti)
    
    try:
        print "about to save best classifier as classifier.pck"
        moMyArgs.Dump
        best.moMyArgs = moMyArgs
        best.gdEmoji = TNAWordAnalyser.gdEmoji
        dump(best, "classifier.pck")
    except NameError:
        pass

    try:
        test = makeTFIDF(lTestData)
        sFileName = "E-C_" + LANGUAGE.lower() + "_pred.txt"
        codaformat(test, best, sFileName)
    except NameError:
        pass
