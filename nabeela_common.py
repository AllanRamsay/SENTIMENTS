#!/usr/bin/python
# -*- coding: utf-8 -*-

#########################################################
#                                                       #
# Standard shared routines for reading Nabeelas files	#
# 22/08/2017                                            #
#														#
#########################################################

import os, sys
from openpyxl import load_workbook
from enum import Enum

# next 3 lines are used to access files in other folders
import copy
sys.path.append(os.path.abspath(os.path.join('', 'Modules')))
import LSA.cNLPLibrary
from LSA.ANRVStemmer import *	# English 4 way stemmer
from colored import fg, bg, attr

global moRed ; 	moRed = fg("#ff0000") + bg("#ffffff")
global moGreen ; moGreen = fg("#00ff00") + bg("#ffffff")
global moBlue ; moBlue = fg("#0000ff") + bg("#ffffff")
global moBlack ; moBlack = fg("#000000") + bg("#ffffff")
global moGold ; moGold = fg("#ffd700") + bg("#ffffff")
global moPurple ; moPurple = fg("magenta") + bg("#ffffff")

global SPECIAL_CHARS ; SPECIAL_CHARS = [u"’", u",", u";", u":", u"!", u"¡", u"?", u"¿", u".", u"..", u"...", u".....", u"„", u"@", u"/", u"\\", u"%", u"†", u"‡", u"´", u"˜", u"¯", u"¨", u"¸", u"ˆ", u"¬", u"¦", u"¢", u"$", u"£", u"¹", u"½", u"¹ï", u"³", u"¾", u"ª", u"â", u"ð", u"ï", u"œ", u"š", u"ÿ", u"ž"]
global mlMandatoryWords ; mlMandatoryWords = None

global mdcDataSets ; mdcDataSets = {}
# NABEELA-12EMOTIONS	Nabeela file with 12 emotions
# NABEELA-4EMOTIONS		Nabeela file with only the 4 same emotions as Saif uses
mdcDataSets['NABEELA-AR-12EMOTIONS'] = "Data/Nabeela/tweets_/16-08-17.xlsx"
mdcDataSets['NABEELA-AR-4EMOTIONS'] = "Data/Nabeela/tweets_/16-08-17-AFJS.xlsx"
mdcDataSets['SAIF-EN-WASSA2017'] = "Data/SAIF/WASSA_2017_EN/DATA.xlsx"
mdcDataSets['SAIF-EN-WASSA2017-TEST'] = "Data/SAIF/WASSA_2017_EN/DATA-TEST.xlsx"
mdcDataSets['SAIF-AR-SYRIAN'] = "Data/SAIF/Syrian_Tweets_AR/DATA.xlsx"
mdcDataSets['SAIF-AR-BBN'] = "Data/SAIF/BBN_Blog_Posts/DATA.xlsx"
mdcDataSets['SAIF-AR-SEMEVAL2018-EMOTION'] = "Data/SAIF/semeval_2018_AR/saifsemeval2018.xlsx"
mdcDataSets['SAIF-AR-SEMEVAL2018-CLASSIFICATION-AR'] = "Data/SAIF/SEMEVAL_2018_CLASSIFICATION/Data/AR/DATA_SEMEVAL2018_RAW_AR.xlsx_formatted.xlsx"
mdcDataSets['SAIF-AR-SEMEVAL2018-CLASSIFICATION-EN'] = "Data/SAIF/SEMEVAL_2018_CLASSIFICATION/Data/EN/DATA_SEMEVAL2018_RAW_EN.xlsx_formatted.xlsx"
mdcDataSets['SAIF-AR-SEMEVAL2018-CLASSIFICATION-ES'] = "Data/SAIF/SEMEVAL_2018_CLASSIFICATION/Data/ES/DATA_SEMEVAL2018_RAW_ES.xlsx_formatted.xlsx"

class MyArgsBase:
	Dedupe = True
	ClassifyOn = None
	Dataset = None
	Folds = None
	PRINT_DEV = False
	Limit = None
	PKeys = None
	Stemmer = None
	Emotion2Mode = None
	IgnoreFirstNRows = None
	KillCache = False
	DataSource = None
	LowerCase = None
	ClusterMode = None
	RemovePunctuationAR = False
	RemovePunctuationEN = False
	RemoveARVowels = False
	RemoveEnglish = False
	RemoveNumbersAR = False
	RemoveNumbersEN = False
	PrepositionsFile = None
	HashtagProcessMode = None
	RemoveAtTags = False
	RemoveURLs = False
	OutputMode = "BASIC"
	TFIDFMode = None
	TweetWordCountRange = None
	RemoveSpecialChars = None
	TestIDs = None
	MandatoryWordsFile = None
	Multipliers = None
	Filter = None	# a filter for columns in excel
	Classification = None
	EmotionList = None	# Internally set. True if tweets per emotion are a list (e.g. from SemEval 2018)
	SemevalThreshold = None	# Only applicable when using a SEMEVAL dataset
	SemevalTuning = None	# Only applicable when using a SEMEVAL dataset

class Classification:
	Scores = None
	Classified = None
	ClassifiedEmotion = None

class Multipliers:
	Emoji = None
	Noun = None
	Verb = None

class TrainingData:
	ID = 0					# internal ID generated in code
	ExternalID = None		# external ID as in the spreadsheet
	Emotion = None
	Emotion2 = None
	Polarity = None
	RawWords = None
	Vector = None
	Tweet = None
	RawTweet = None
	OriginalEmotion = None	# original emotion before clustering applied
	OriginalEmotion2 = None	# original emotion before clustering applied
	OriginalTweet = None	# the original tweet sent to NA with emojis etc
	TestTweet = None		# tweet is explicitly flagged as a test tweet
	Tag = None				# user for internal stuff
	Data = None				# the data to work with

class ClusterMode(Enum):
	Humaine = 1

class EmotionPair:
	Emotion1 = None
	Emotion2 = None

def AddTestItem(oMyArgs, oTrainingData):
	bAdd=False
	if oMyArgs.TestIDs==None:
		bAdd=True
	else:
		for sTestID in oMyArgs.TestIDs:
			if sTestID.upper() == oTrainingData.ID.upper():
				bAdd = True
		
	return bAdd

def ExpandTweetHashtags(sTweet):
	asHashtags = LSA.cNLPLibrary.ExtractHashtags(sTweet)
	for sHashtag in asHashtags:
		sHashtag = sHashtag.replace("#","").replace("_"," ").replace("-"," ").strip()
		sTweet = sTweet + " " + sHashtag
	sTweet = sTweet.strip()
	return sTweet

def ExtractPlaceHolder(sFilter):
	sRes = (sFilter.split("["))[1].split("]")[0]
	return sRes

def GetHumaineClusters():
	lRes = []
	oEmotionPair = EmotionPair() ; oEmotionPair.Emotion1 = 'anger' ; oEmotionPair.Emotion2 = 'disgust' ; lRes.append(oEmotionPair)
	oEmotionPair = EmotionPair() ; oEmotionPair.Emotion1 = 'sad' ; oEmotionPair.Emotion2 = 'disappointment' ; lRes.append(oEmotionPair)
	oEmotionPair = EmotionPair() ; oEmotionPair.Emotion1 = 'sad' ; oEmotionPair.Emotion2 = 'bored' ; lRes.append(oEmotionPair)
	oEmotionPair = EmotionPair() ; oEmotionPair.Emotion1 = 'satisfied' ; oEmotionPair.Emotion2 = 'trust' ; lRes.append(oEmotionPair)
	oEmotionPair = EmotionPair() ; oEmotionPair.Emotion1 = 'shame' ; oEmotionPair.Emotion2 = 'pride' ; lRes.append(oEmotionPair)
	return lRes

def GetClusteredEmotion(lClusters, sEmotion2):
	sRes = None
	for oEmotionPair in lClusters:
		if oEmotionPair.Emotion2 == sEmotion2:
			sRes = oEmotionPair.Emotion1
			break
	return sRes

def LoadArgs(oArgs, oMyArgs):
	global mlMandatoryWords
	global mdcDataSets

	bOK = False

	for sArg in oArgs:
		sArg = sArg.upper()
		if (sArg.startswith("DEDUPE=")):
			if str(sArg.split("=")[1]).upper()=="N":
				oMyArgs.Dedupe=False
		if (sArg.startswith("CLASSIFYON=")):
			oMyArgs.ClassifyOn = str(sArg.split("=")[1])
		if (sArg.startswith("DATASET=")):
			oMyArgs.Dataset = str(sArg.split("=")[1])
		if (sArg.startswith("EMOTION2MODE=")):
			oMyArgs.Emotion2Mode = str(sArg.split("=")[1])
		if (sArg.startswith("FOLDS=")):
			oMyArgs.Folds = str(sArg.split("=")[1])
		if (sArg.startswith("PRINT_DEV=")):
			if str(sArg.split("=")[1]).upper()=="Y":
				oMyArgs.PRINT_DEV=True
		if (sArg.startswith("LIMIT=")):
			oMyArgs.Limit = int(sArg.split("=")[1])
		if (sArg.startswith("PKEYS=")):
			oMyArgs.PKeys = str(sArg.split("=")[1])
		if (sArg.startswith("STEMMER=")):
			oMyArgs.Stemmer = int(sArg.split("=")[1])
		if (sArg.startswith("IGNOREFIRST=")):
			oMyArgs.IgnoreFirstNRows = int(sArg.split("=")[1])
		if (sArg.startswith("KILLCACHE")):
			oMyArgs.KillCache=True
		if (sArg.startswith("DATASOURCE=")):
			oMyArgs.DataSource = str(sArg.split("=")[1])
		if (sArg.startswith("CLUSTERMODE=")):
			sClusterMode = str(sArg.split("=")[1])
			if sClusterMode.upper()=="HUMAINE":
				oMyArgs.ClusterMode = ClusterMode.Humaine
		if (sArg.startswith("REMOVEPUNCTUATIONAR")):
			oMyArgs.RemovePunctuationAR=True
		if (sArg.startswith("REMOVEPUNCTUATIONEN")):
			oMyArgs.RemovePunctuationEN=True
		if (sArg.startswith("REMOVEARVOWELS")):
			oMyArgs.RemoveARVowels=True
		if (sArg.startswith("REMOVEENGLISH")):
			oMyArgs.RemoveEnglish=True
		if (sArg.startswith("REMOVENUMBERSAR")):
			oMyArgs.RemoveNumbersAR=True
		if (sArg.startswith("REMOVENUMBERSEN")):
			oMyArgs.RemoveNumbersEN=True
		if (sArg.startswith("PREPOSITIONSFILE")):
			oMyArgs.PrepositionsFile = str(sArg.split("=")[1])
		if (sArg.startswith("MANDATORYWORDSFILE")):
			oMyArgs.MandatoryWordsFile = str(sArg.split("=")[1])
			mlMandatoryWords = LoadFile('Resources/' + oMyArgs.MandatoryWordsFile, mlMandatoryWords)
		if (sArg.startswith("HASHTAGPROCESSMODE")):
			oMyArgs.HashtagProcessMode = str(sArg.split("=")[1])
		if (sArg.startswith("REMOVEATTAGS")):
			oMyArgs.RemoveAtTags=True
		if (sArg.startswith("REMOVEURLS")):
			oMyArgs.RemoveURLs=True
		if (sArg.startswith("OUTPUTMODE")):
			oMyArgs.OutputMode=str(sArg.split("=")[1])
		if (sArg.startswith("TFIDFMODE=")):
			oMyArgs.TFIDFMode = int(sArg.split("=")[1])
		if (sArg.startswith("TWEETWORDCOUNTRANGE=")):
			oMyArgs.TweetWordCountRange = str(sArg.split("=")[1])
		if (sArg.startswith("LOWERCASE")):
			oMyArgs.LowerCase=True
		if (sArg.startswith("REMOVESPECIALCHARS")):
			oMyArgs.RemoveSpecialChars=True
		if (sArg.startswith("SEMEVALTHRESHOLD=")):
			oMyArgs.SemevalThreshold = float(sArg.split("=")[1])
		if (sArg.startswith("SEMEVALTUNING=")):
			oMyArgs.SemevalTuning = int(sArg.split("=")[1])
		if (sArg.startswith("TESTIDS")):
			# test specific items only
			oMyArgs.TestIDs = str(sArg.split("=")[1]).split(",")
		if (sArg.startswith("FILTER")):
			oMyArgs.Filter = str(sArg.split("=")[1]).split(",")
		if (sArg.startswith("MULTIPLIERS")):
			#example MULTIPLIERS=E:10;N:9;V:5
			sMultipliers = str(sArg.split("=")[1])
			asMultipliers = sMultipliers.split(";")
			oMultipliers = Multipliers()
			for sMultiplier in asMultipliers:
				sKey = sMultiplier.split(":")[0]
				nValue = int(sMultiplier.split(":")[1])
				if sKey=="E": oMultipliers.Emoji = nValue
				if sKey=="N": oMultipliers.Noun = nValue
				if sKey=="V": oMultipliers.Verb = nValue
			oMyArgs.Multipliers = oMultipliers

	# REMOVEHASHTAGS			remove the complete hashtag
	# REMOVE#					remove just the # part of the hashtag
	# COPY_REMOVE#_SPLIT		copy the original hashtags, remove the # part of the hashtag, split the rest into words
	# REMOVE#_SPLIT				remove the # part of the hashtag, split the rest into words
	lHashtagProcessModes = ['REMOVEHASHTAGS','REMOVE#','COPY_REMOVE#_SPLIT','REMOVE#_SPLIT']

	# BASIC			default
	# WEKA			for WEKA input
	# VERBOSE		for detailed output
	# CODALAB2017	for input into CODALAB 2017 evaluation.py
	# CODALAB20171	for input into CODALAB 2017 evaluation.py with my confidence scores 1
	# CODALAB20172	for input into CODALAB 2017 evaluation.py with my confidence scores 2
	lOutputModes = ['BASIC','WEKA','VERBOSE','CODALAB2017','CODALAB20171','CODALAB20172']

	# 0		No stemming
	# 1		AR version 1 PAR stemmer
	# 2		AR version 2 PAR stemmer
	# 3		EN Morphy+Wordnet stemmer
	# 4		EN ARNV 4 way take-the-smallest stemmer
	lStemmers = [0,1,2,3,4]

	if oMyArgs.ClassifyOn==None:
		print "ERROR: No CLASSIFYON specified (EMOTION | POLARITY)"
	elif oMyArgs.ClassifyOn!="EMOTION" and oMyArgs.ClassifyOn!="POLARITY":
		print "ERROR: Incorrect CLASSIFYON (EMOTION or POLARITY) specified"	
	elif ((oMyArgs.HashtagProcessMode!=None) and (not oMyArgs.HashtagProcessMode in lHashtagProcessModes)):
		print "ERROR: Incorrect HASHTAGPROCESSMODE specified:",lHashtagProcessModes
	elif ((oMyArgs.OutputMode!=None) and (not oMyArgs.OutputMode in lOutputModes)):
		print "ERROR: Incorrect OUTPUTMODE specified:",lOutputModes
	elif oMyArgs.Dataset==None:
		print "ERROR: No DATASET specified"
	elif ((oMyArgs.Dataset!=None) and (not oMyArgs.Dataset in mdcDataSets.keys())):
		print "ERROR: Incorrect DATASET specified:", mdcDataSets.keys()
	elif oMyArgs.TFIDFMode==None:
		print "ERROR: No TFIDFMODE specified"
	elif oMyArgs.IgnoreFirstNRows==None:
		print "ERROR: No IGNOREFIRST specified"
	elif oMyArgs.Stemmer==None:
		print "ERROR: No STEMMER specified"
	elif (not oMyArgs.Stemmer in lStemmers):
		print "ERROR: Incorrect STEMMER specified:", lStemmers
	elif oMyArgs.Folds==None:
		print "ERROR: No FOLDS specified"
	elif oMyArgs.DataSource==None:
		print "ERROR: No DATASOURCE (TWEETS or WORDS) specified"
	elif oMyArgs.DataSource!="TWEETS" and oMyArgs.DataSource!="WORDS" and oMyArgs.DataSource!="OTWEETS":
		print "ERROR: Incorrect DATASOURCE (TWEETS or WORDS or OTWEETS) specified"
	else:
		bOK = True
		if oMyArgs.Dataset.find("SEMEVAL")>-1:
			if oMyArgs.SemevalThreshold==None:
				print "ERROR: SEMEVAL dataset requires a SEMEVALTHRESHOLD"
				bOK = False

	return bOK

def LoadFile(sFile, lList):

	if lList == None:
		with io.open(sFile, 'r', encoding='utf-8', errors='replace') as f:
			lList = f.readlines()
			lList = [oItem.replace("\n","") for oItem in lList] 

	return lList

def LoadTrainingData(lTrainingData, oMyArgs, sEmotionCell1, sEmotionCell2, sPolarityCell, sWordsCell, sTweetCell, sTimeCell, sIDCell, sOriginalTweetCell, sTestFlagCell, bDedupe, eClusterMode=None):
	global SPECIAL_CHARS

	sFilePath = mdcDataSets[oMyArgs.Dataset]

	sDir, sFile =  os.path.split(sFilePath)
	oWB = load_workbook(sFilePath)
	asColumns = [sEmotionCell1, sEmotionCell2, sPolarityCell, sWordsCell, sTweetCell, sOriginalTweetCell, sTestFlagCell, sIDCell]

	print "Loading training data from: ", sFilePath

	# A workbook is always created with at least one worksheet
	oWS = oWB.active

	nRow=0
	nPhysicalRows=0
	nBlankTweet=0
	nBlankEmotion=0
	nDuplicates=0
	nNA=0
	nMultiEmotionRows=0
	nNTweetLenOutsideRange=0
	nNoContainMandatoryWords=0
	lAllRows = []
	nTweetWordCountFrom=None
	nTweetWordCountTo=None
	nFailedFilter=0

	if oMyArgs.TweetWordCountRange!=None:
		nTweetWordCountFrom = int(oMyArgs.TweetWordCountRange.split("-")[0])
		nTweetWordCountTo = int(oMyArgs.TweetWordCountRange.split("-")[1])

	if eClusterMode == ClusterMode.Humaine:
		lClusters = GetHumaineClusters()

	for oRow in oWS.iter_rows():
		nRow+=1

		if oMyArgs.Limit != None:
			if len(lTrainingData)>=oMyArgs.Limit:
				break

		# ignore the header row
		if nRow>1 and nRow>oMyArgs.IgnoreFirstNRows:
			# create an object for saving the data
			oTrainingData = TrainingData()
			# used for dedupe purposes
			sRowData = ""
			for oCell in oRow:
				sColumn = oCell.column
				sCellName = "{}{}".format(sColumn, nRow)
				sValue = oWS[sCellName].value # the value of the specific cell

				# build string of all cell values for dedupe
				if sValue!=None:
					if not sColumn in [sTimeCell, sIDCell]: 
						if type(sValue) in ['float']:
							sRowData = sRowData + str(sValue)
						else:
							sRowData = sRowData + sValue

				if sColumn in asColumns:
					if sValue!=None:
						if type(sValue) not in ["float", "long"]:
							sValue = sValue.strip()
					# save it
					if sColumn == sIDCell:
						oTrainingData.ExternalID = str(sValue)
					if sColumn == sEmotionCell1:
						oTrainingData.Emotion = sValue
						oTrainingData.OriginalEmotion = sValue
						if sValue!=None:
							if sValue.find(",")>-1:
								# flag the emotions as being a list
								oMyArgs.EmotionList = True
					if sColumn == sEmotionCell2:
						oTrainingData.Emotion2 = sValue
						oTrainingData.OriginalEmotion2 = sValue
					if sColumn == sPolarityCell:
						oTrainingData.Polarity = sValue
					if sColumn == sTweetCell:
						if oMyArgs.LowerCase==True:
							if sValue!=None:
								sValue = sValue.lower()
						if oMyArgs.RemoveSpecialChars==True:
							if sValue!=None:
								for sC in SPECIAL_CHARS:
									sValue = sValue.replace(sC, "")
						oTrainingData.Tweet = NormalizeArabic(sValue)
					if sColumn == sWordsCell:
						if oMyArgs.LowerCase==True:
							if sValue!=None:
								sValue = sValue.lower()
						if oMyArgs.RemoveSpecialChars==True:
							if sValue!=None:
								for sC in SPECIAL_CHARS:
									sValue = sValue.replace(sC,"")
						oTrainingData.RawWords = NormalizeArabic(sValue) ###########
					if sColumn == sOriginalTweetCell:
						oTrainingData.RawTweet = sValue
						if oMyArgs.LowerCase==True:
							if sValue!=None:
								sValue = sValue.lower()
						if oMyArgs.RemoveSpecialChars==True:
							if sValue!=None:
								for sC in SPECIAL_CHARS:
									sValue = sValue.replace(sC,"")
						oTrainingData.OriginalTweet = NormalizeArabic(sValue) #########
					if sColumn == sTestFlagCell:
						if sValue!=None:
							if sValue!="":
								oTrainingData.TestTweet = True
				# apply any clustering
				if eClusterMode != None:
					if oTrainingData.Emotion!="" and oTrainingData.Emotion!=None:
						sClusteredEmotion = GetClusteredEmotion(lClusters, oTrainingData.Emotion)
						if sClusteredEmotion!=None:
							oTrainingData.Emotion = sClusteredEmotion
					if oTrainingData.Emotion2!="" and oTrainingData.Emotion2!=None:
						sClusteredEmotion = GetClusteredEmotion(lClusters, oTrainingData.Emotion2)
						if sClusteredEmotion!=None:
							oTrainingData.Emotion2 = sClusteredEmotion

			# if classifying on Polarity set the emotion to the Polarity
			if oMyArgs.ClassifyOn=="POLARITY":
				oTrainingData.Emotion = oTrainingData.Polarity
				# same row, multiple emotions, but they will both have the same Polarity as they are on the same row 
				oTrainingData.Emotion2 = ""

			if bDedupe:
				bOK = True
				if sRowData in lAllRows:
					nDuplicates+=1
					bOK = False
			else:
				bOK = True

			lsTweetWords = None

			if bOK:
				if nTweetWordCountFrom!=None and nTweetWordCountTo!=None:
					if oTrainingData.OriginalTweet != "" and oTrainingData.OriginalTweet != None:
						if lsTweetWords==None:
							lsTweetWords = NLTKTokeniser2.casual_tokenize(oTrainingData.OriginalTweet)
						nLen = len(lsTweetWords)
						if nLen<nTweetWordCountFrom or nLen>nTweetWordCountTo:
							bOK = False
							nNTweetLenOutsideRange+=1
			
			if bOK:
				if oTrainingData.OriginalTweet == "" or oTrainingData.OriginalTweet == None:
					bOK = False

			if bOK:
				if mlMandatoryWords!=None:
					# see if the tweet contains any of the words specified
					if lsTweetWords==None:
						lsTweetWords = NLTKTokeniser2.casual_tokenize(oTrainingData.OriginalTweet)
					bOK = False
					for sWord in lsTweetWords:
						if sWord in mlMandatoryWords:
							bOK = True
							break
					
					if bOK == False:
						nNoContainMandatoryWords+=1

			if bOK:
				if oMyArgs.Filter != None:
					sFilter = oMyArgs.Filter[0]
					# replace any commands, we cant use = in args 
					sFilter = sFilter.replace("GE", ">=")
					sFilter = sFilter.replace("LE", "<=")
					sFilter = sFilter.replace("GT", ">")
					sFilter = sFilter.replace("LT", "<")
					sCol =  ExtractPlaceHolder(sFilter)
					sCellName = "{}{}".format(sCol, nRow)
					sValue = oWS[sCellName].value # the value of the specific cell
					sFilter = sFilter.replace("[" + sCol + "]", str(sValue))
					bOK = eval(sFilter)

					if bOK==False:
						nFailedFilter+=1

			if bOK:
				lAllRows.append(sRowData)
				if oTrainingData.Tweet != "" and oTrainingData.Tweet != None:

					bOK = False

					if oMyArgs.ClassifyOn=="EMOTION":
						if oTrainingData.Emotion!=None and oTrainingData.Emotion!="":
							if oTrainingData.Emotion.upper()!="NA":
								bOK = True
							else:
								nNA+=1
						else:
							nBlankEmotion+=1

					if oMyArgs.ClassifyOn=="POLARITY":
						if oTrainingData.Polarity!=None and oTrainingData.Polarity!="":
							if oTrainingData.Polarity.upper()=="POSITIVE" or oTrainingData.Polarity.upper()=="NEGATIVE":
								bOK = True
							else:
								nNA+=1
						else:
							nBlankEmotion+=1

					if bOK:
						oTrainingData.ID = sFile + "_" + str(nRow)
						lTrainingData.append(oTrainingData)
						# if we have a second emotion then duplicate the record (depending on mode)
						if oMyArgs.Emotion2Mode!="IGNORE":
							if oTrainingData.Emotion2!="" and oTrainingData.Emotion2!=None:
								if oTrainingData.Emotion2.upper()!="NA":
									nMultiEmotionRows+=1
									oTrainingData2 = copy.copy(oTrainingData)
									sEmotion1 = oTrainingData.Emotion
									sEmotion2 = oTrainingData.Emotion2
									oTrainingData2.Emotion = sEmotion2
									oTrainingData2.Emotion2 = sEmotion1
									oTrainingData2.ID = sFile + "_" + str(nRow) + "_E2"
									lTrainingData.append(oTrainingData2)
				else:
					nBlankTweet+=1

	bRemoveHashtags = False
	bRemoveHashtagSymbol = False
	bCopyRemoveSplitHashtag = False
	bRemoveHashtagSymbolSplit = False

	if oMyArgs.HashtagProcessMode!=None or \
		oMyArgs.RemoveAtTags or \
		oMyArgs.RemoveURLs or \
		oMyArgs.RemoveNumbersAR or \
		oMyArgs.RemoveNumbersEN or \
		oMyArgs.RemoveEnglish or \
		oMyArgs.RemovePunctuationEN or \
		oMyArgs.PrepositionsFile!=None or \
		oMyArgs.RemovePunctuationAR or \
		oMyArgs.RemoveARVowels:

		for oData in lTrainingData:

			if oMyArgs.HashtagProcessMode == "REMOVEHASHTAGS":
				bRemoveHashtags = True

			if oMyArgs.HashtagProcessMode == "REMOVE#":
				bRemoveHashtagSymbol = True

			if oMyArgs.HashtagProcessMode == "COPY_REMOVE#_SPLIT":
				bCopyRemoveSplitHashtag = True

			if oMyArgs.HashtagProcessMode == "REMOVE#_SPLIT":
				bRemoveHashtagSymbolSplit = True

			if bCopyRemoveSplitHashtag==True:
				oData.Tweet = ExpandTweetHashtags(oData.Tweet)
				if oData.RawWords!=None: oData.RawWords = ExpandTweetHashtags(oData.RawWords)
				oData.OriginalTweet = ExpandTweetHashtags(oData.OriginalTweet)

			if bRemoveHashtagSymbol==True:
				if oData.RawWords!=None: oData.RawWords = oData.RawWords.replace("#", "")
				if oData.Tweet!=None: oData.Tweet = oData.Tweet.replace("#", "")
				if oData.OriginalTweet!=None: oData.OriginalTweet = oData.OriginalTweet.replace("#", "")

			if bRemoveHashtagSymbolSplit==True:
				if oData.RawWords!=None: oData.RawWords = oData.RawWords.replace("#","").replace("_"," ").replace("-"," ").strip()
				if oData.Tweet!=None: oData.Tweet = oData.Tweet.replace("#","").replace("_"," ").replace("-"," ").strip()
				if oData.OriginalTweet!=None: oData.OriginalTweet = oData.OriginalTweet.replace("#","").replace("_"," ").replace("-"," ").strip()

			if oData.RawWords!=None:
				oData.RawWords = RemoveTokensByClass(oData.RawWords, bRemoveHashtags, False, oMyArgs.RemoveAtTags, oMyArgs.RemoveURLs)
			oData.OriginalTweet = RemoveTokensByClass(oData.OriginalTweet, bRemoveHashtags, False, oMyArgs.RemoveAtTags, oMyArgs.RemoveURLs)
			oData.Tweet = RemoveTokensByClass(oData.Tweet, bRemoveHashtags, False, oMyArgs.RemoveAtTags, oMyArgs.RemoveURLs)

			if oMyArgs.RemoveNumbersAR==True:
				oData.RawWords = RemoveNumbersAR(oData.RawWords)
				oData.Tweet = RemoveNumbersAR(oData.Tweet)
				oData.OriginalTweet = RemoveNumbersAR(oData.OriginalTweet)

			if oMyArgs.RemoveNumbersEN==True:
				oData.RawWords = RemoveNumbersEN(oData.RawWords)
				oData.Tweet = RemoveNumbersEN(oData.Tweet)
				oData.OriginalTweet = RemoveNumbersEN(oData.OriginalTweet)

			if oMyArgs.RemoveEnglish==True:
				oData.RawWords = RemoveEnglish(oData.RawWords)
				oData.Tweet = RemoveEnglish(oData.Tweet)
				oData.OriginalTweet = RemoveEnglish(oData.OriginalTweet)

			if oMyArgs.RemovePunctuationEN==True:
				if oData.RawWords!=None:
					oData.RawWords = RemovePunctuationCharactersEN(oData.RawWords)
				oData.Tweet = RemovePunctuationCharactersEN(oData.Tweet)
				oData.OriginalTweet = RemovePunctuationCharactersEN(oData.OriginalTweet)

			if oMyArgs.PrepositionsFile!=None:
				if oData.RawWords!=None:
					oData.RawWords = RemovePrepositions(oData.RawWords, oMyArgs.PrepositionsFile)
				oData.Tweet = RemovePrepositions(oData.Tweet, oMyArgs.PrepositionsFile)
				oData.OriginalTweet = RemovePrepositions(oData.OriginalTweet, oMyArgs.PrepositionsFile)

			if oMyArgs.RemovePunctuationAR==True:
				if oData.RawWords!=None:
					oData.RawWords = RemovePunctuationCharactersAR(oData.RawWords)
				oData.Tweet = RemovePunctuationCharactersAR(oData.Tweet)
				oData.OriginalTweet = RemovePunctuationCharactersAR(oData.OriginalTweet)
			
			if oMyArgs.RemoveARVowels==True:
				if oData.RawWords!=None:
					oData.RawWords = RemoveHarakaat(oData.RawWords)
				oData.Tweet = RemoveHarakaat(oData.Tweet)
				oData.OriginalTweet = RemoveHarakaat(oData.OriginalTweet)

	# set the .Data property for ease
	for oTrainingData in lTrainingData:
		if oMyArgs.DataSource=="OTWEETS":
			oTrainingData.Data = oTrainingData.OriginalTweet
		if oMyArgs.DataSource=="TWEETS":
			oTrainingData.Data = oTrainingData.Tweet
		if oMyArgs.DataSource=="WORDS":
			oTrainingData.Data = oTrainingData.RawWords

	if oMyArgs.EmotionList:
		print moRed
		print "EMOTIONLIST mode is ON"
		print moBlack
		# convert all emotions into a list
		for oTrainingData in lTrainingData:
			oTrainingData.Emotion = oTrainingData.Emotion.split(",")

	# show summary
	print "Rows in file", nRow
	print "Duplicate rows", nDuplicates
	print "Rows with no tweet", nBlankTweet
	print "Rows with no emotion", nBlankEmotion
	print "Rows with NA emotion", nNA
	print "Rows with multiple emotions", nMultiEmotionRows
	print "Tweet length outside specified range", nNTweetLenOutsideRange
	print "Tweets that dont contain mandatory words", nNoContainMandatoryWords
	print "Failed to pass Excel filter clause", nFailedFilter
	print "#Total tweets (across all files)", len(lTrainingData)
	print ""

def NormaliseDictionaryValuesTo(dc, lfNormaliseTo):
	lfFactor=lfNormaliseTo/sum(dc.itervalues())
	for k in dc:
		dc[k] = dc[k]*lfFactor
	return dc

def PrintArgs(oMyArgs, bRecursing=False):
	for attr in dir(oMyArgs):
		if not attr.startswith("__"):
			if bRecursing:
				print "   ",
			print("%s = %s" % (attr, getattr(oMyArgs, attr)))
			if isinstance(getattr(oMyArgs, attr), Multipliers):
				PrintArgs(getattr(oMyArgs, attr), True)

	if not bRecursing:
		print ""

def SplitData(lAllData, oMyArgs):
	# now split into test and training
	# the function will return a list of dictionaries
	# each item in the list is a dictionary.
	# the dictionary will contain 2 items, a test set and a training set
	print "Creating cross folds"

	nBadData = 0
	nPotSwapped = 0
	bOverrideTestFlag = False
	bTestDataFlagged = TestDataFlagged(lAllData)

	try:
		nFolds = int(oMyArgs.Folds)
		bFoldsIsInt = True
	except:
		bFoldsIsInt = False

	if not bFoldsIsInt:
		# the FOLDS parameter is (any) non-integer indicating that we 
		# we wish to use test data flagged in the file
		nFolds = 1
		if not bTestDataFlagged:
			print "ERROR: No test data found in file."
			exit()
		else:
			print "Using test data found in file."
	else:
		nFolds = int(oMyArgs.Folds)
		# if we have specified folds then cannot use the test data
		bTestDataFlagged = False

	# create a dictionary of the fold size
	# the dictionary will be keyed on fold number (eg FOLD3)
	# each element will be a dictionary, which will be list of test
	# and training items keyed on TEST and TRAIN
	dcRes = {}

	for nFold in range(1,nFolds+1):
		dcFold={}
		dcFold['TEST'] = []
		dcFold['TRAIN'] = []
		dcRes['FOLD' + str(nFold)] = dcFold

	if bTestDataFlagged:

		for oTrainingData in lAllData:
			# create 1 fold based on flag in database
			if oTrainingData.TestTweet:
				sKey = "TEST"
				bAdd = AddTestItem(oMyArgs, oTrainingData)
			else:
				sKey = "TRAIN"
				bAdd = True
			if bAdd:
				dcRes['FOLD1'][sKey].append(oTrainingData)

	else:

		# create N folds
		nIndex = 1
		nCurrentFold=1
		for oTrainingData in lAllData:
			bOK = False
			if oMyArgs.DataSource=="WORDS":
				bOK = oTrainingData.RawWords!=None and oTrainingData.RawWords!=""
			if oMyArgs.DataSource=="TWEETS":
				bOK = oTrainingData.Tweet!=None and oTrainingData.Tweet!=""
			if oMyArgs.DataSource=="OTWEETS":
				bOK = oTrainingData.Tweet!=None and oTrainingData.Tweet!=""

			if bOK:
				# each item needs to go into each of the dictionaries
				# either as a test item or as training item
				for nFold in range(1,nFolds+1):
					oCopy = copy.copy(oTrainingData)
					oCopy.ID = nIndex###################DELETE#######################
					if (nFold == nCurrentFold) or (nPotSwapped>0):
						sKey = "TEST"
						# cannot have multiple emotions for test data
						if oTrainingData.Emotion2!=None and oTrainingData.Emotion2!="":
							sKey="TRAIN"
							if (nFold == nCurrentFold):
								nPotSwapped+=1
							oCopy.ID = str(oCopy.ID) + str("POTSWAP") ###################DELETE#######################
						else:
							nPotSwapped=nPotSwapped-1
							if nPotSwapped<0: nPotSwapped=0
					else:
						sKey = "TRAIN"

					bAdd = True
					if sKey=="TEST":
						bAdd = AddTestItem(oMyArgs, oTrainingData)	

					if bAdd:
						dcRes['FOLD' + str(nFold)][sKey].append(oCopy)

			else:
				nBadData+=1
			
			nCurrentFold+=1
			nIndex+=1
			if nCurrentFold>nFolds:
				nCurrentFold = 1

	print "#Pot swapped", nPotSwapped
	print "#Missing Tweet or RawWords", nBadData

	print "\n\n-=-=-=-=-=-=-=-=-=-=-=-FOLD SIZES-=-=-=-=-=-=-=-=-=-=-=-"
	for nFold in range(1,nFolds+1):

		dcFold = dcRes['FOLD' + str(nFold)]
		nTrain = len(dcFold['TRAIN'])
		nTest = len(dcFold['TEST'])
		print "Fold ", nFold, ": Train=", nTrain, " Test=", nTest, (float(nTest)/(float(nTest)+float(nTrain))) * 100 

	return dcRes

"""
def SplitDataOLD(lAllData, oMyArgs, lTrainingData, lTestData):
	# now split into test and training
	nIndex=0
	nBadData=0
	nPotSwapped = 0

	# check if the file has any tweets specifically flagged as test data
	bTestDataFlagged = TestDataFlagged(lAllData)

	for oTrainingData in lAllData:
		nIndex+=1
		sPot = ""
		bOK = False

		if bTestDataFlagged:
			if oTrainingData.TestTweet:
				if AddTestItem(oMyArgs, oTrainingData):
					lTestData.append(oTrainingData)
			else:
				lTrainingData.append(oTrainingData)
		else:
			if oMyArgs.DataSource=="WORDS":
				bOK = oTrainingData.RawWords!=None and oTrainingData.RawWords!=""
			if oMyArgs.DataSource=="TWEETS":
				bOK = oTrainingData.Tweet!=None and oTrainingData.Tweet!=""
			if oMyArgs.DataSource=="OTWEETS":
				bOK = oTrainingData.Tweet!=None and oTrainingData.Tweet!=""

			if bOK:
				bTestTweet = ((nIndex % oMyArgs.Folds)==0)
				if (bTestTweet or nPotSwapped>0):
					sPot="TEST"
					# cannot have multiple emotions for test data
					if oTrainingData.Emotion2!=None and oTrainingData.Emotion2!="":
						sPot="TRAIN"
						if bTestTweet:
							nPotSwapped+=1
					else:
						nPotSwapped=nPotSwapped-1
						if nPotSwapped<0: nPotSwapped=0
				else:
					sPot="TRAIN"

				if sPot=="TRAIN":
					lTrainingData.append(oTrainingData)
				if sPot=="TEST":
					if AddTestItem(oMyArgs, oTrainingData):
						lTestData.append(oTrainingData)
			else:
				nBadData+=1
	
	return lTrainingData, lTestData, nBadData, nIndex
"""

def TagStemTrainingData(lTrainingData, oMyArgs):

	# AR flavour
	if oMyArgs.Stemmer == 1 or oMyArgs.Stemmer == 2:

		# tagger takes ages to instantiate so do all tagging in one go
		oI = FABTaggerHelper()
		oI.Clear()
		for oTrainingData in lTrainingData:
			sTweet = oTrainingData.Data
			oI.Append(sTweet)
		
		oTaggedResults = oI.Tag()
		
		# now stem the tagged results using old and new tagger and set the oTrainingData property
		nIndex=0
		for sTaggedTweet in oTaggedResults:
			sTaggedTweet = cLibrary.StripNewline(sTaggedTweet)

			if oMyArgs.Stemmer == 1:
				oS = FABStemmerARHelper()
				oS.Clear()
				oS.Append(sTaggedTweet)
				sStemmedTweet = oS.Stem(False)
				sStemmedTweet = LSA.cNLPLibrary.ReduceStemmedTagged2(sStemmedTweet, ['NN','VV'])
				lTrainingData[nIndex].Data = sStemmedTweet

			if oMyArgs.Stemmer == 2:
				oS2 = FABStemmer2Helper()
				oS2.Clear()
				oS2.Append(sTaggedTweet)
				sStemmedTweet = oS2.Stem(False)
				sStemmedTweet = LSA.cNLPLibrary.ReduceStemmedTagged2(sStemmedTweet, ['NAME','NSTEM','VSTEM'])
				lTrainingData[nIndex].Data = sStemmedTweet

			nIndex+=1

	# EN flavour
	if oMyArgs.Stemmer == 3:
		nIndex=0
		for oTrainingData in lTrainingData:
			sTweet = oTrainingData.Tweet
			oS = MorphyStemmerHelper()
			oS.Clear()
			oS.Append(sTweet)
			sTweet = oS.Stem()
			lTrainingData[nIndex].Data = sTweet
			nIndex+=1

	if oMyArgs.Stemmer == 4:
		nIndex=0
		for oTrainingData in lTrainingData:
			sTweet = oTrainingData.Data
			oS = ARNVStemmerHelper()
			oS.Clear()
			oS.Append(sTweet)
			sTweet = oS.StemARNV()
			lTrainingData[nIndex].Data = sTweet
			nIndex+=1

def TestDataFlagged(lData):
	bRes = False

	for oData in lData:
		if oData.TestTweet!=None and oData.TestTweet!="": 
			bRes = True
			break

	return bRes
